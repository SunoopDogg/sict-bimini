"""xlsx → list[BIMObjectRaw] parser (openpyxl-based).

The xlsx layout (after the required header row) follows this convention:

- ``객체 유형: <IfcXxx>`` — sticky type declaration; applies to every
  subsequent object until a new declaration overrides it.
- ``GlobalId: <id>`` — object boundary; finalizes the previous object
  (if any) and starts a new one.
- Any other row — property row for the current object.

Ported from legacy ``sict-bimini-python/src/converters/xlsx2json.py`` with:
- pandas/NaN removed → openpyxl + ``None`` tracking
- returns typed ``BIMObjectRaw`` instead of raw dicts
"""

from __future__ import annotations

import logging
from pathlib import Path
from typing import TypedDict

from openpyxl import load_workbook

from api.bim.schemas import BIMObjectRaw

logger = logging.getLogger(__name__)

REQUIRED_COLUMNS: tuple[str, ...] = ("객체명", "속성세트", "속성명", "속성값")


class MissingColumnsError(ValueError):
    """Required columns missing from the xlsx header."""


class _Current(TypedDict):
    global_id: str | None
    object_name: str
    properties: dict[str, dict[str, str]]
    ifc_type: str | None


def _cell(value: object) -> str | None:
    """Normalize a cell value: ``None`` stays ``None``, others become str."""
    if value is None:
        return None
    return str(value)


def _is_type_declaration(obj_name: str) -> bool:
    return obj_name.startswith("객체유형") or obj_name.startswith("객체 유형")


def _is_global_id(obj_name: str) -> bool:
    return obj_name.lower().startswith("globalid")


def _split_after_colon(text: str) -> str:
    return text.split(":", 1)[1].strip() if ":" in text else ""


def parse_xlsx_to_raw(path: Path) -> list[BIMObjectRaw]:
    """Parse a BIM xlsx file into a list of :class:`BIMObjectRaw`."""
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active

    try:
        header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    except StopIteration:
        return []

    header_index = {name: i for i, name in enumerate(header) if name}
    missing = [c for c in REQUIRED_COLUMNS if c not in header_index]
    if missing:
        raise MissingColumnsError(
            f"Missing required columns: {missing}. Found: {list(header)}"
        )

    col_obj = header_index["객체명"]
    col_set = header_index["속성세트"]
    col_prop = header_index["속성명"]
    col_val = header_index["속성값"]

    # openpyxl read_only truncates trailing None cells; force padding.
    row_iter = ws.iter_rows(min_row=2, max_col=len(header), values_only=True)

    source_file = path.name
    bim_objects: list[BIMObjectRaw] = []
    current: _Current | None = None
    ifc_type: str | None = None  # sticky across objects

    for row in row_iter:
        obj_name = _cell(row[col_obj])
        prop_set = _cell(row[col_set])
        prop_name = _cell(row[col_prop])
        prop_val = _cell(row[col_val])

        if not any((obj_name, prop_set, prop_name, prop_val)):
            continue

        if obj_name and _is_type_declaration(obj_name):
            after = _split_after_colon(obj_name)
            if after:
                ifc_type = after if after.startswith("Ifc") else f"Ifc{after}"
            continue

        if obj_name and _is_global_id(obj_name):
            if current is not None:
                bim_objects.append(_finalize(current, source_file))
            current = {
                "global_id": _split_after_colon(obj_name) or None,
                "object_name": "",
                "properties": {},
                "ifc_type": ifc_type,
            }
            continue

        if current is None:
            continue

        if not current["object_name"] and obj_name:
            current["object_name"] = obj_name

        if prop_set is not None and prop_name is not None:
            bucket = current["properties"].setdefault(prop_set, {})
            bucket[prop_name] = prop_val if prop_val is not None else ""

    if current is not None:
        bim_objects.append(_finalize(current, source_file))

    logger.info("Parsed %d objects from %s", len(bim_objects), source_file)
    return bim_objects


def _finalize(current: _Current, source_file: str) -> BIMObjectRaw:
    return BIMObjectRaw(
        source_file=source_file,
        object_name=current["object_name"],
        ifc_type=current["ifc_type"],
        global_id=current["global_id"],
        properties=current["properties"],
    )
